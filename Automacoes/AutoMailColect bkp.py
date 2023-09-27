import win32com.client
import re
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time
import os
from datetime import datetime
from dateutil import parser

# Buscas no título e corpo do email e tempo de intervalo
palavra_titulo_en = "PIX deposit"
palavra_titulo_br = "Depósito PIX"

palavra_user_en = "account username origin:"
palavra_user_br = "Nome de usuário  origem:"

palavra_valor_en = "Deposit amount:"
palavra_valor_br = "Quantia depositada:"

tempo_intervalo = 10

data_de_hoje = datetime.now().strftime('%d/%m/%Y')

excel_file = 'mail.xlsx'
uspus1 = '00331-10000-00001-AA281'

if win32com.client.Dispatch("WScript.Shell").RegRead("HKLM\SOFTWARE\Microsoft\Windows NT\CurrentVersion\ProductID") == uspus1:
    outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
    inbox = outlook.GetDefaultFolder(6)
    print(inbox)

    def extract_name_and_value(line, palavra_user, palavra_valor):
        name_match = re.search(fr'{palavra_user}\s*(.*)', line)
        value_match = re.search(fr'{palavra_valor}\s*([\d,.]+)', line)
        name = name_match.group(1).strip() if name_match else ''
        value = value_match.group(1).strip() if value_match else ''
        return name, value

    total_novos_emails = sum(1 for item in inbox.Items if item.Class == 43 and item.UnRead and 
                            (palavra_titulo_en in item.Subject or palavra_titulo_br in item.Subject))

    if total_novos_emails > 0:
        print(f'Total de novos emails encontrados: {total_novos_emails}')

        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['OP', 'Data', 'Username', 'Ação(DEP/Withdraw)', 'Valor', 'Descrição', 'Observação Extra'])
            for cell in ws['D1:E1']:
                for column_cell in cell:
                    column_cell.fill = PatternFill(start_color='93c47d', end_color='93c47d', fill_type='solid')

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
                for cell in row:
                    cell.value = data_de_hoje

        emails_ids_coletados = []

        for item in inbox.Items:
            if item.Class == 43 and item.UnRead and (palavra_titulo_en in item.Subject or palavra_titulo_br in item.Subject):
                email_id = item.EntryID
                if email_id not in emails_ids_coletados:
                    emails_ids_coletados.append(email_id)
                    email_text = item.Body
                    lines = email_text.split('\n')
                    nome = ''
                    valor = ''
                    data_hora = item.SentOn.strftime('%d/%m/%Y %H:%M:%S')
                    for line in lines:
                        line = line.strip()
                        if palavra_user_en in line or palavra_user_br in line:
                            nome, _ = extract_name_and_value(line, palavra_user_en if palavra_user_en in line else palavra_user_br, 
                                                            palavra_valor_en if palavra_user_en in line else palavra_valor_br)
                        elif palavra_valor_en in line or palavra_valor_br in line:
                            _, valor = extract_name_and_value(line, palavra_user_en if palavra_valor_en in line else palavra_user_br, 
                                                             palavra_valor_en if palavra_valor_en in line else palavra_valor_br)
                    if nome and valor:
                        emails_coletados = [['', data_de_hoje, nome, 'Dep', valor, '', '']]
                        item.UnRead = False
                        print(f'Assunto do email: {item.Subject} do usuário: {nome} e com o valor: {valor}')
                        print(f"Aguardando {tempo_intervalo} segundos para o próximo email\n")
                        emails_coletados.sort(key=lambda x: parser.parse(x[1]))
                        for email_data in emails_coletados:
                            ws.append(email_data)

                        for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=4, max_col=5):
                            for cell in row:
                                cell.fill = PatternFill(start_color='93c47d', end_color='93c47d', fill_type='solid')

                        wb.save(excel_file)
                        time.sleep(tempo_intervalo)
        print(f'Dados dos emails salvos em: {excel_file}')
    else:
        print('Nenhum novo email correspondente aos critérios encontrado.')
else:
    print('Usuário não possui as permissões para usar o código.')